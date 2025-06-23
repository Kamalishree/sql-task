from openpyxl import Workbook, load_workbook

#  Create a workbook and worksheet
file_path = r"C:\Users\Asus\Downloads\Sample3.xlsx"  # <-- Adjust path if needed

wb = Workbook()
sheet = wb.active
sheet.title = "Employees"

# Add headers
sheet["A1"] = "Emp ID"
sheet["B1"] = "Name"
sheet["C1"] = "Department"

# Add data
sheet.append([101, "Kamali", "IT"])
sheet.append([102, "Ravi", "HR"])
sheet.append([103, "Meena", "Finance"])

#  Save workbook
wb.save(file_path)

# Read and print data
wb = load_workbook(file_path)
sheet = wb["Employees"]

print("Employee Records:")
for row in sheet.iter_rows(min_row=2, values_only=True):
    emp_id, name, dept = row
    print(f"ID: {emp_id}, Name: {name}, Department: {dept}")
